"""Excel export service with multi-tab support."""
import io
from datetime import datetime
from typing import Any, Dict, List, Optional
from uuid import UUID

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.entity import Entity, EntityBatch, EntityOwnership, EntityResolution
import structlog

logger = structlog.get_logger()


class ExportService:
    """Service for exporting data to Excel with multiple tabs."""
    
    # Styling constants
    HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    SUBHEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ALT_ROW_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    THIN_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def __init__(self, db: AsyncSession):
        self.db = db
    
    async def export_batch_to_excel(
        self,
        batch_id: UUID,
        include_resolutions: bool = True,
        include_ownership: bool = True,
        include_financial: bool = True,
        include_enriched: bool = True,
    ) -> bytes:
        """
        Export batch data to multi-tab Excel file.
        
        Args:
            batch_id: Batch to export
            include_resolutions: Include resolution candidates tab
            include_ownership: Include ownership tree tab
            include_financial: Include financial summary tab
            include_enriched: Include enriched data tab
        
        Returns:
            Excel file as bytes
        """
        # Get batch with entities
        result = await self.db.execute(
            select(EntityBatch)
            .options(selectinload(EntityBatch.entities))
            .where(EntityBatch.id == batch_id)
        )
        batch = result.scalar_one_or_none()
        
        if not batch:
            raise ValueError(f"Batch {batch_id} not found")
        
        # Get all entities
        entities_result = await self.db.execute(
            select(Entity)
            .where(Entity.batch_id == batch_id)
            .order_by(Entity.row_number)
        )
        entities = entities_result.scalars().all()
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        
        # Tab 1: Summary
        await self._create_summary_sheet(wb, batch, entities)
        
        # Tab 2: All Entities
        await self._create_entities_sheet(wb, entities)
        
        # Tab 3: Resolution Candidates
        if include_resolutions:
            await self._create_resolutions_sheet(wb, entities)
        
        # Tab 4: Ownership Tree
        if include_ownership:
            await self._create_ownership_sheet(wb, entities)
        
        # Tab 5: Financial Data
        if include_financial:
            await self._create_financial_sheet(wb, entities)
        
        # Tab 6: Enriched Data
        if include_enriched:
            await self._create_enriched_sheet(wb, entities)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()
    
    async def _create_summary_sheet(
        self,
        wb: Workbook,
        batch: EntityBatch,
        entities: List[Entity],
    ):
        """Create summary overview sheet."""
        ws = wb.create_sheet("Summary")
        
        # Title
        ws.merge_cells('A1:D1')
        ws['A1'] = f"Charity Data Enrichment Report: {batch.name}"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Batch info
        ws['A3'] = "Batch Information"
        ws['A3'].font = Font(bold=True, size=12)
        
        info = [
            ("Batch ID:", str(batch.id)),
            ("Batch Name:", batch.name),
            ("Original File:", batch.original_filename),
            ("Created:", batch.created_at.strftime("%Y-%m-%d %H:%M:%S") if batch.created_at else "N/A"),
            ("Processed:", batch.processing_completed_at.strftime("%Y-%m-%d %H:%M:%S") if batch.processing_completed_at else "N/A"),
            ("Status:", batch.status.value if batch.status else "Unknown"),
        ]
        
        for i, (label, value) in enumerate(info, start=4):
            ws[f'A{i}'] = label
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'] = value
        
        # Statistics
        ws['A12'] = "Statistics"
        ws['A12'].font = Font(bold=True, size=12)
        
        matched = sum(1 for e in entities if e.resolution_status.value == "matched")
        confirmed = sum(1 for e in entities if e.resolution_status.value == "confirmed")
        no_match = sum(1 for e in entities if e.resolution_status.value == "no_match")
        pending = sum(1 for e in entities if e.resolution_status.value == "pending")
        review = sum(1 for e in entities if e.resolution_status.value in ("multiple_matches", "manual_review"))
        
        stats = [
            ("Total Records:", len(entities)),
            ("Matched:", matched),
            ("Confirmed:", confirmed),
            ("No Match:", no_match),
            ("Pending Review:", review),
            ("Pending:", pending),
            ("Match Rate:", f"{((matched + confirmed) / len(entities) * 100):.1f}%" if entities else "0%"),
        ]
        
        for i, (label, value) in enumerate(stats, start=13):
            ws[f'A{i}'] = label
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'B{i}'] = value
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50
    
    async def _create_entities_sheet(self, wb: Workbook, entities: List[Entity]):
        """Create main entities data sheet."""
        ws = wb.create_sheet("Entities")
        
        # Prepare data
        data = []
        for entity in entities:
            data.append({
                "Row #": entity.row_number,
                "Original Name": entity.original_name,
                "Resolved Name": entity.resolved_name,
                "Entity Type": entity.entity_type.value if entity.entity_type else "",
                "Charity Number": entity.charity_number,
                "Company Number": entity.company_number,
                "Status": entity.charity_status,
                "Resolution Status": entity.resolution_status.value if entity.resolution_status else "",
                "Confidence": f"{entity.resolution_confidence:.1%}" if entity.resolution_confidence else "",
                "Method": entity.resolution_method,
                "Registration Date": entity.charity_registration_date.strftime("%Y-%m-%d") if entity.charity_registration_date else "",
                "Website": entity.charity_website,
                "Email": entity.charity_contact_email,
                "Address": entity.charity_address,
            })
        
        if not data:
            ws['A1'] = "No entities found"
            return
        
        df = pd.DataFrame(data)
        
        # Write headers
        for col_idx, column in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.THIN_BORDER
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for row_idx, row in enumerate(df.values, start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.THIN_BORDER
                if row_idx % 2 == 0:
                    cell.fill = self.ALT_ROW_FILL
        
        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    async def _create_resolutions_sheet(self, wb: Workbook, entities: List[Entity]):
        """Create resolution candidates sheet."""
        ws = wb.create_sheet("Resolution Candidates")
        
        # Get all resolutions
        entity_ids = [e.id for e in entities]
        result = await self.db.execute(
            select(EntityResolution)
            .where(EntityResolution.entity_id.in_(entity_ids))
            .order_by(EntityResolution.entity_id, EntityResolution.confidence_score.desc())
        )
        resolutions = result.scalars().all()
        
        # Create entity lookup
        entity_lookup = {e.id: e for e in entities}
        
        # Prepare data
        data = []
        for res in resolutions:
            entity = entity_lookup.get(res.entity_id)
            data.append({
                "Original Name": entity.original_name if entity else "",
                "Candidate Name": res.candidate_name,
                "Charity Number": res.charity_number,
                "Confidence Score": f"{res.confidence_score:.1%}",
                "Match Method": res.match_method,
                "Selected": "Yes" if res.is_selected else "No",
            })
        
        if not data:
            ws['A1'] = "No resolution candidates found"
            return
        
        df = pd.DataFrame(data)
        
        # Write headers
        for col_idx, column in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.THIN_BORDER
        
        # Write data
        for row_idx, row in enumerate(df.values, start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.THIN_BORDER
                if row_idx % 2 == 0:
                    cell.fill = self.ALT_ROW_FILL
        
        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    async def _create_ownership_sheet(self, wb: Workbook, entities: List[Entity]):
        """Create ownership tree sheet."""
        ws = wb.create_sheet("Ownership Tree")
        
        # Get ownership relationships
        entity_ids = [e.id for e in entities]
        result = await self.db.execute(
            select(EntityOwnership)
            .where(
                (EntityOwnership.owner_id.in_(entity_ids)) |
                (EntityOwnership.owned_id.in_(entity_ids))
            )
        )
        ownerships = result.scalars().all()
        
        # Create entity lookup
        entity_lookup = {e.id: e for e in entities}
        
        # Prepare data
        data = []
        for ownership in ownerships:
            owner = entity_lookup.get(ownership.owner_id)
            owned = entity_lookup.get(ownership.owned_id)
            
            data.append({
                "Owner Name": owner.resolved_name or owner.original_name if owner else "Unknown",
                "Owner Charity #": owner.charity_number if owner else "",
                "Relationship": ownership.ownership_type,
                "Owned Entity": owned.resolved_name or owned.original_name if owned else "Unknown",
                "Owned Charity #": owned.charity_number if owned else "",
                "Owned Company #": owned.company_number if owned else "",
                "Ownership %": f"{ownership.ownership_percentage:.1f}%" if ownership.ownership_percentage else "",
                "Source": ownership.source,
                "Verified": "Yes" if ownership.verified else "No",
            })
        
        if not data:
            ws['A1'] = "No ownership relationships found"
            return
        
        df = pd.DataFrame(data)
        
        # Write headers
        for col_idx, column in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.THIN_BORDER
        
        # Write data
        for row_idx, row in enumerate(df.values, start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.THIN_BORDER
                if row_idx % 2 == 0:
                    cell.fill = self.ALT_ROW_FILL
        
        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    async def _create_financial_sheet(self, wb: Workbook, entities: List[Entity]):
        """Create financial summary sheet."""
        ws = wb.create_sheet("Financial Data")
        
        # Prepare data
        data = []
        for entity in entities:
            if entity.latest_income or entity.latest_expenditure:
                data.append({
                    "Name": entity.resolved_name or entity.original_name,
                    "Charity Number": entity.charity_number,
                    "Status": entity.charity_status,
                    "Latest Income": entity.latest_income,
                    "Latest Expenditure": entity.latest_expenditure,
                    "Net Position": (entity.latest_income or 0) - (entity.latest_expenditure or 0),
                    "Financial Year End": entity.latest_financial_year_end.strftime("%Y-%m-%d") if entity.latest_financial_year_end else "",
                })
        
        if not data:
            ws['A1'] = "No financial data available"
            return
        
        df = pd.DataFrame(data)
        
        # Write headers
        for col_idx, column in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.border = self.THIN_BORDER
        
        # Write data
        for row_idx, row in enumerate(df.values, start=2):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.THIN_BORDER
                if row_idx % 2 == 0:
                    cell.fill = self.ALT_ROW_FILL
                # Format currency columns
                if col_idx in (4, 5, 6) and isinstance(value, (int, float)):
                    cell.number_format = '£#,##0.00'
        
        # Summary row
        total_row = len(data) + 3
        ws[f'A{total_row}'] = "TOTALS"
        ws[f'A{total_row}'].font = Font(bold=True)
        
        total_income = sum(d.get("Latest Income") or 0 for d in data)
        total_expenditure = sum(d.get("Latest Expenditure") or 0 for d in data)
        
        ws[f'D{total_row}'] = total_income
        ws[f'D{total_row}'].number_format = '£#,##0.00'
        ws[f'D{total_row}'].font = Font(bold=True)
        
        ws[f'E{total_row}'] = total_expenditure
        ws[f'E{total_row}'].number_format = '£#,##0.00'
        ws[f'E{total_row}'].font = Font(bold=True)
        
        ws[f'F{total_row}'] = total_income - total_expenditure
        ws[f'F{total_row}'].number_format = '£#,##0.00'
        ws[f'F{total_row}'].font = Font(bold=True)
        
        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    async def _create_enriched_sheet(self, wb: Workbook, entities: List[Entity]):
        """Create enriched data sheet with trustees and subsidiaries."""
        ws = wb.create_sheet("Enriched Data")
        
        # Prepare trustees data
        trustees_data = []
        subsidiaries_data = []
        
        for entity in entities:
            if entity.enriched_data:
                trustees = entity.enriched_data.get("trustees", [])
                for trustee in trustees:
                    trustees_data.append({
                        "Charity Name": entity.resolved_name or entity.original_name,
                        "Charity Number": entity.charity_number,
                        "Trustee Name": trustee.get("name", ""),
                        "Trustee ID": trustee.get("id", ""),
                    })
                
                subsidiaries = entity.enriched_data.get("subsidiaries", [])
                for sub in subsidiaries:
                    subsidiaries_data.append({
                        "Charity Name": entity.resolved_name or entity.original_name,
                        "Charity Number": entity.charity_number,
                        "Subsidiary Name": sub.get("name", ""),
                        "Company Number": sub.get("company_number", ""),
                    })
        
        # Write Trustees section
        ws['A1'] = "TRUSTEES"
        ws['A1'].font = Font(bold=True, size=14)
        
        if trustees_data:
            df = pd.DataFrame(trustees_data)
            
            for col_idx, column in enumerate(df.columns, start=1):
                cell = ws.cell(row=2, column=col_idx, value=column)
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT
                cell.border = self.THIN_BORDER
            
            for row_idx, row in enumerate(df.values, start=3):
                for col_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = self.THIN_BORDER
                    if row_idx % 2 == 0:
                        cell.fill = self.ALT_ROW_FILL
            
            next_row = len(trustees_data) + 5
        else:
            ws['A3'] = "No trustee data available"
            next_row = 6
        
        # Write Subsidiaries section
        ws[f'A{next_row}'] = "SUBSIDIARIES"
        ws[f'A{next_row}'].font = Font(bold=True, size=14)
        
        if subsidiaries_data:
            df = pd.DataFrame(subsidiaries_data)
            
            for col_idx, column in enumerate(df.columns, start=1):
                cell = ws.cell(row=next_row + 1, column=col_idx, value=column)
                cell.fill = self.SUBHEADER_FILL
                cell.font = self.HEADER_FONT
                cell.border = self.THIN_BORDER
            
            for row_idx, row in enumerate(df.values, start=next_row + 2):
                for col_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = self.THIN_BORDER
                    if row_idx % 2 == 0:
                        cell.fill = self.ALT_ROW_FILL
        else:
            ws[f'A{next_row + 2}'] = "No subsidiary data available"
        
        # Auto-fit columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    async def export_to_csv(self, batch_id: UUID) -> bytes:
        """Export basic entity data to CSV."""
        result = await self.db.execute(
            select(Entity)
            .where(Entity.batch_id == batch_id)
            .order_by(Entity.row_number)
        )
        entities = result.scalars().all()
        
        data = []
        for entity in entities:
            data.append({
                "row_number": entity.row_number,
                "original_name": entity.original_name,
                "resolved_name": entity.resolved_name,
                "entity_type": entity.entity_type.value if entity.entity_type else "",
                "charity_number": entity.charity_number,
                "company_number": entity.company_number,
                "status": entity.charity_status,
                "resolution_status": entity.resolution_status.value if entity.resolution_status else "",
                "confidence": entity.resolution_confidence,
                "income": entity.latest_income,
                "expenditure": entity.latest_expenditure,
                "website": entity.charity_website,
                "email": entity.charity_contact_email,
            })
        
        df = pd.DataFrame(data)
        output = io.BytesIO()
        df.to_csv(output, index=False)
        output.seek(0)
        return output.getvalue()
